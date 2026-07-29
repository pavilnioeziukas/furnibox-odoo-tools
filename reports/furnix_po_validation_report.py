from __future__ import annotations

from collections import Counter
from datetime import date, datetime, time, timezone
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl.styles import Font, PatternFill

from core.bom_engine import BomEngine
from core.bom_repository import BomRepository
from core.bom_selector import BomSelector
from core.odoo_helpers import many2one_id, many2one_name
from core.report_base import BaseReport
from validators.furnix_po_validator import (
    FurnixPoValidationResult,
    FurnixPoValidator,
)


LITHUANIA_TIMEZONE = ZoneInfo("Europe/Vilnius")
FURNIX_VENDOR_NAME = "FURNIX, UAB"

MODE_UNCONFIRMED = "unconfirmed"
MODE_APPROVED = "approved"
MODE_TODAY_SO = "today_so"


class FurnixPoValidationReport(BaseReport):
    name = "Furnix PO patikra"
    description = (
        "Patikrina Furnix PO detales, kiekius ir "
        "Component Sticker Info prieš arba po PO patvirtinimo."
    )

    def __init__(self, client: Any) -> None:
        super().__init__(client)

        repository = BomRepository(client)
        selector = BomSelector(repository)
        engine = BomEngine(repository, selector)

        self.validator = FurnixPoValidator(
            client,
            engine,
        )

    def run(self) -> None:
        while True:
            print("\n====================================")
            print("FURNIX PO PATIKRA")
            print("====================================")
            print("1. PO paruošti tvirtinimui")
            print("2. Patikrinti jau patvirtintus PO")
            print("3. Patikrinti šiandien sukurtus SO")
            print("0. Grįžti")

            choice = input(
                "\nPasirinkite patikros būdą: "
            ).strip()

            if choice == "0":
                return

            if choice == "1":
                self.run_mode(MODE_UNCONFIRMED)
                return

            if choice == "2":
                self.run_mode(MODE_APPROVED)
                return

            if choice == "3":
                self.run_mode(MODE_TODAY_SO)
                return

            print("\nNeteisingas pasirinkimas.")

    def run_mode(self, mode: str) -> None:
        print("\nNuskaitomi duomenys iš Odoo...")

        if mode == MODE_UNCONFIRMED:
            source_records = self.load_unconfirmed_furnix_pos()
            title = "Nepatvirtintų Furnix PO patikra"

        elif mode == MODE_APPROVED:
            source_records = self.load_approved_furnix_pos()
            title = "Patvirtintų Furnix PO auditas"

        elif mode == MODE_TODAY_SO:
            source_records = self.load_todays_sale_orders()
            title = "Šiandien sukurtų SO patikra"

        else:
            raise ValueError(
                f"Nežinomas patikros režimas: {mode}"
            )

        print(f"\n{title}")
        print(f"Rasta įrašų: {len(source_records)}")

        if not source_records:
            print("\nTikrintinų įrašų nerasta.")
            return

        results = self.validate_records(
            mode=mode,
            records=source_records,
        )

        summary_dataframe = self.build_summary_dataframe(
            results,
            mode=mode,
        )

        details_dataframe = self.build_details_dataframe(
            results
        )

        self.print_summary(
            summary_dataframe,
            mode=mode,
        )

        output_path = self.export_to_excel(
            summary_dataframe=summary_dataframe,
            details_dataframe=details_dataframe,
            mode=mode,
        )

        print("\nAtaskaita sukurta:")
        print(output_path)

    def load_unconfirmed_furnix_pos(
        self,
    ) -> list[dict[str, Any]]:
        purchase_orders = self.client.search_read(
            model="purchase.order",
            domain=[
                ["state", "not in", ["purchase", "done", "cancel"]],
            ],
            fields=[
                "name",
                "state",
                "partner_id",
                "sale_primary_id",
                "create_date",
            ],
            order="create_date asc, id asc",
        )

        return self.filter_furnix_purchase_orders(
            purchase_orders
        )

    def load_approved_furnix_pos(
        self,
    ) -> list[dict[str, Any]]:
        purchase_orders = self.client.search_read(
            model="purchase.order",
            domain=[
                ["state", "in", ["purchase", "done"]],
            ],
            fields=[
                "name",
                "state",
                "partner_id",
                "sale_primary_id",
                "create_date",
            ],
            order="create_date desc, id desc",
            limit=50,
        )

        return self.filter_furnix_purchase_orders(
            purchase_orders
        )

    def load_todays_sale_orders(
        self,
    ) -> list[dict[str, Any]]:
        target_date = datetime.now(
            LITHUANIA_TIMEZONE
        ).date()

        start_local = datetime.combine(
            target_date,
            time.min,
            tzinfo=LITHUANIA_TIMEZONE,
        )

        end_local = datetime.combine(
            target_date,
            time.max,
            tzinfo=LITHUANIA_TIMEZONE,
        )

        start_utc = start_local.astimezone(
            timezone.utc
        )

        end_utc = end_local.astimezone(
            timezone.utc
        )

        return self.client.search_read(
            model="sale.order",
            domain=[
                [
                    "create_date",
                    ">=",
                    start_utc.strftime(
                        "%Y-%m-%d %H:%M:%S"
                    ),
                ],
                [
                    "create_date",
                    "<=",
                    end_utc.strftime(
                        "%Y-%m-%d %H:%M:%S"
                    ),
                ],
            ],
            fields=[
                "name",
                "state",
                "partner_id",
                "create_date",
            ],
            order="create_date asc, id asc",
        )

    @staticmethod
    def filter_furnix_purchase_orders(
        purchase_orders: list[dict[str, Any]],
    ) -> list[dict[str, Any]]:
        return [
            purchase_order
            for purchase_order in purchase_orders
            if many2one_name(
                purchase_order.get("partner_id")
            ).strip().upper()
            == FURNIX_VENDOR_NAME
        ]

    def validate_records(
        self,
        *,
        mode: str,
        records: list[dict[str, Any]],
    ) -> list[
        tuple[
            dict[str, Any],
            FurnixPoValidationResult,
        ]
    ]:
        results: list[
            tuple[
                dict[str, Any],
                FurnixPoValidationResult,
            ]
        ] = []

        for index, record in enumerate(
            records,
            start=1,
        ):
            if mode == MODE_TODAY_SO:
                so_number = str(
                    record.get("name") or ""
                )
            else:
                so_number = many2one_name(
                    record.get("sale_primary_id")
                )

            print(
                f"\rTikrinama {index}/{len(records)}: "
                f"{so_number or '-':<25}",
                end="",
                flush=True,
            )

            if not so_number:
                result = FurnixPoValidationResult(
                    so_number="",
                    po_number=str(
                        record.get("name") or ""
                    ),
                    po_state=str(
                        record.get("state") or ""
                    ),
                    vendor_name=many2one_name(
                        record.get("partner_id")
                    ),
                    furnix_po_count=1,
                    status="BOM ERROR",
                    error=(
                        "PO neturi užpildyto "
                        "Primary Sale Order."
                    ),
                )
            else:
                result = self.validator.validate(
                    so_number
                )

            results.append(
                (
                    record,
                    result,
                )
            )

        print()

        return results

    def build_summary_dataframe(
        self,
        results: list[
            tuple[
                dict[str, Any],
                FurnixPoValidationResult,
            ]
        ],
        *,
        mode: str,
    ) -> pd.DataFrame:
        rows: list[dict[str, Any]] = []

        for source_record, result in results:
            is_purchase_order_source = (
                mode != MODE_TODAY_SO
            )

            source_po = (
                str(source_record.get("name") or "")
                if is_purchase_order_source
                else ""
            )

            source_po_state = (
                str(source_record.get("state") or "")
                if is_purchase_order_source
                else ""
            )

            rows.append(
                {
                    "SO": result.so_number,
                    "PO": result.po_number or source_po,
                    "PO State": (
                        result.po_state
                        or source_po_state
                    ),
                    "Vendor": (
                        result.vendor_name
                        or many2one_name(
                            source_record.get(
                                "partner_id"
                            )
                        )
                    ),
                    "Furnix PO Count": (
                        result.furnix_po_count
                    ),
                    "Status": result.status,
                    "PASS Lines": result.pass_count,
                    "Missing": result.missing_count,
                    "Extra": result.extra_count,
                    "Qty Mismatch": (
                        result.qty_mismatch_count
                    ),
                    "Sticker Errors": (
                        result.sticker_error_count
                    ),
                    "Action": self.action_for_result(
                        status=result.status,
                        po_state=(
                            result.po_state
                            or source_po_state
                        ),
                        mode=mode,
                    ),
                    "Error": result.error,
                }
            )

        return pd.DataFrame(rows)

    @staticmethod
    def build_details_dataframe(
        results: list[
            tuple[
                dict[str, Any],
                FurnixPoValidationResult,
            ]
        ],
    ) -> pd.DataFrame:
        rows: list[dict[str, Any]] = []

        for _, result in results:
            for row in result.rows:
                origins = "; ".join(
                    (
                        f"{origin.get('group_name', '')}: "
                        f"{origin.get('root_sku', '')} "
                        f"[QTY:{origin.get('required_qty', '')}]"
                    )
                    for origin in row.origins
                )

                rows.append(
                    {
                        "SO": result.so_number,
                        "PO": result.po_number,
                        "PO State": result.po_state,
                        "SKU": row.sku,
                        "Product": row.product_name,
                        "Required Qty": row.required_qty,
                        "PO Qty": row.po_qty,
                        "Difference": row.difference,
                        "BOM / SO Origin": origins,
                        "Component Sticker Info": " | ".join(
                            row.component_sticker_info
                        ),
                        "Sticker Status": (
                            row.sticker_status
                        ),
                        "Row Status": row.status,
                        "Sticker Error": (
                            row.sticker_error
                        ),
                    }
                )

        return pd.DataFrame(rows)

    @staticmethod
    def action_for_result(
        *,
        status: str,
        po_state: str,
        mode: str,
    ) -> str:
        is_approved = po_state in {
            "purchase",
            "done",
        }

        if status == "PASS":
            if is_approved:
                return "NO ACTION"

            return "CONFIRM PO"

        if is_approved or mode == MODE_APPROVED:
            return "URGENT - CHECK / CONTACT SUPPLIER"

        actions = {
            "NO PO": "WAIT / CHECK PO",
            "MULTIPLE PO": "STOP - CHECK DUPLICATES",
            "MISSING": "STOP - MISSING DETAILS",
            "EXTRA": "STOP - EXTRA DETAILS",
            "QTY MISMATCH": (
                "STOP - CHECK QUANTITIES"
            ),
            "STICKER INFO ERROR": (
                "STOP - FIX STICKER INFO"
            ),
            "BOM ERROR": "STOP - CHECK BOM",
        }

        return actions.get(
            status,
            "STOP - CHECK",
        )

    @staticmethod
    def print_summary(
        dataframe: pd.DataFrame,
        *,
        mode: str,
    ) -> None:
        print("\nPATIKROS SUVESTINĖ")
        print("=" * 70)

        status_counts = Counter(
            dataframe["Status"].fillna("")
        )

        for status, count in sorted(
            status_counts.items()
        ):
            print(
                f"{status or 'UNKNOWN':<25}"
                f"{count}"
            )

        confirmable = dataframe[
            dataframe["Action"] == "CONFIRM PO"
        ]

        urgent = dataframe[
            dataframe["Action"].str.startswith(
                "URGENT",
                na=False,
            )
        ]

        blocked = dataframe[
            dataframe["Action"].str.startswith(
                "STOP",
                na=False,
            )
        ]

        print("\nGALIMA TVIRTINTI:")
        if confirmable.empty:
            print("- nėra")
        else:
            for _, row in confirmable.iterrows():
                print(
                    f"- {row['SO']} / {row['PO']}"
                )

        if mode == MODE_APPROVED:
            print("\nSKUBIAI TIKRINTI:")
            if urgent.empty:
                print("- nėra")
            else:
                for _, row in urgent.iterrows():
                    print(
                        f"- {row['SO']} / "
                        f"{row['PO']}: "
                        f"{row['Status']}"
                    )
        else:
            print("\nNEGALIMA TVIRTINTI:")
            if blocked.empty:
                print("- nėra")
            else:
                for _, row in blocked.iterrows():
                    print(
                        f"- {row['SO']} / "
                        f"{row['PO'] or '-'}: "
                        f"{row['Status']}"
                    )

    @staticmethod
    def export_to_excel(
        *,
        summary_dataframe: pd.DataFrame,
        details_dataframe: pd.DataFrame,
        mode: str,
    ) -> Path:
        output_directory = Path("output")
        output_directory.mkdir(
            parents=True,
            exist_ok=True,
        )

        today = date.today().isoformat()

        output_path = (
            output_directory
            / f"Furnix_PO_Validation_{mode}_{today}.xlsx"
        )

        with pd.ExcelWriter(
            output_path,
            engine="openpyxl",
        ) as writer:
            summary_dataframe.to_excel(
                writer,
                index=False,
                sheet_name="Summary",
            )

            if details_dataframe.empty:
                pd.DataFrame(
                    columns=[
                        "SO",
                        "PO",
                        "PO State",
                        "SKU",
                        "Product",
                        "Required Qty",
                        "PO Qty",
                        "Difference",
                        "BOM / SO Origin",
                        "Component Sticker Info",
                        "Sticker Status",
                        "Row Status",
                        "Sticker Error",
                    ]
                ).to_excel(
                    writer,
                    index=False,
                    sheet_name="Details",
                )
            else:
                details_dataframe.to_excel(
                    writer,
                    index=False,
                    sheet_name="Details",
                )

            summary_sheet = writer.sheets["Summary"]
            details_sheet = writer.sheets["Details"]

            for worksheet in (
                summary_sheet,
                details_sheet,
            ):
                worksheet.freeze_panes = "A2"
                worksheet.auto_filter.ref = (
                    worksheet.dimensions
                )

                for cell in worksheet[1]:
                    cell.font = Font(bold=True)

            summary_widths = [
                22,
                16,
                14,
                24,
                18,
                24,
                14,
                12,
                12,
                16,
                16,
                38,
                80,
            ]

            for index, width in enumerate(
                summary_widths,
                start=1,
            ):
                summary_sheet.column_dimensions[
                    summary_sheet.cell(
                        row=1,
                        column=index,
                    ).column_letter
                ].width = width

            details_widths = [
                22,
                16,
                14,
                48,
                60,
                14,
                14,
                14,
                80,
                100,
                24,
                20,
                100,
            ]

            for index, width in enumerate(
                details_widths,
                start=1,
            ):
                details_sheet.column_dimensions[
                    details_sheet.cell(
                        row=1,
                        column=index,
                    ).column_letter
                ].width = width

            fills = {
                "PASS": PatternFill(
                    "solid",
                    fgColor="D9EAD3",
                ),
                "NO PO": PatternFill(
                    "solid",
                    fgColor="FFF2CC",
                ),
                "MULTIPLE PO": PatternFill(
                    "solid",
                    fgColor="F4CCCC",
                ),
                "MISSING": PatternFill(
                    "solid",
                    fgColor="F4CCCC",
                ),
                "EXTRA": PatternFill(
                    "solid",
                    fgColor="F4CCCC",
                ),
                "QTY MISMATCH": PatternFill(
                    "solid",
                    fgColor="F4CCCC",
                ),
                "STICKER INFO ERROR": PatternFill(
                    "solid",
                    fgColor="F4CCCC",
                ),
                "BOM ERROR": PatternFill(
                    "solid",
                    fgColor="F4CCCC",
                ),
            }

            status_column = next(
                (
                    cell.column
                    for cell in summary_sheet[1]
                    if cell.value == "Status"
                ),
                None,
            )

            if status_column is not None:
                for row_number in range(
                    2,
                    summary_sheet.max_row + 1,
                ):
                    status = str(
                        summary_sheet.cell(
                            row=row_number,
                            column=status_column,
                        ).value
                        or ""
                    )

                    fill = fills.get(status)

                    if fill is None:
                        continue

                    for column_number in range(
                        1,
                        summary_sheet.max_column + 1,
                    ):
                        summary_sheet.cell(
                            row=row_number,
                            column=column_number,
                        ).fill = fill

        return output_path.resolve()