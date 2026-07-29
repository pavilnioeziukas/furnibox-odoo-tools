from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.config import OdooConfig
from core.odoo_client import OdooClient, OdooConnectionError
from core.odoo_helpers import many2one_id
from core.validated_dataset_repository import (
    ValidatedDatasetRepository,
)


QTY_TOLERANCE = 0.000001
MAX_RECORDS = 100000


def canon(value: Any) -> str:
    return str(value or "").strip().upper()


def to_float(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def parse_write_date(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value

    raw = str(value or "").strip()

    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
    ):
        try:
            return datetime.strptime(
                raw[:19],
                fmt,
            )
        except ValueError:
            continue

    return datetime.min


@dataclass(frozen=True)
class SelectedBom:
    bom_id: int
    parent_sku: str
    reference: str
    bom_type: str
    sequence: int
    product_qty: float
    write_date: str
    duplicate_sequence0_count: int


@dataclass(frozen=True)
class DifferenceRow:
    parent_sku: str
    bom_id: int | None
    bom_reference: str
    status: str
    component_sku: str
    catalog_qty: float
    odoo_qty: float
    difference: float
    duplicate_sequence0_count: int


def load_products(
    client: OdooClient,
) -> tuple[
    dict[int, dict[str, Any]],
    dict[int, list[dict[str, Any]]],
]:
    rows = client.search_read(
        model="product.product",
        domain=[
            ["default_code", "!=", False],
        ],
        fields=[
            "default_code",
            "display_name",
            "product_tmpl_id",
            "active",
        ],
        limit=MAX_RECORDS,
    )

    by_id: dict[int, dict[str, Any]] = {}
    by_template: dict[int, list[dict[str, Any]]] = defaultdict(
        list
    )

    for row in rows:
        product_id = int(row["id"])
        by_id[product_id] = row

        template_id = many2one_id(
            row.get("product_tmpl_id")
        )

        if template_id is not None:
            by_template[template_id].append(
                row
            )

    return by_id, dict(by_template)


def resolve_bom_parent_sku(
    bom: dict[str, Any],
    products_by_id: dict[int, dict[str, Any]],
    products_by_template: dict[int, list[dict[str, Any]]],
) -> str:
    product_id = many2one_id(
        bom.get("product_id")
    )

    if product_id is not None:
        product = products_by_id.get(
            product_id,
            {},
        )

        return canon(
            product.get("default_code")
        )

    template_id = many2one_id(
        bom.get("product_tmpl_id")
    )

    if template_id is None:
        return ""

    candidates = [
        product
        for product in products_by_template.get(
            template_id,
            [],
        )
        if canon(
            product.get("default_code")
        )
    ]

    if len(candidates) == 1:
        return canon(
            candidates[0].get("default_code")
        )

    # Jei template turi kelis variantus, netylime ir nepriskiriame atsitiktinai.
    return ""


def choose_active_sequence0_boms(
    client: OdooClient,
    products_by_id: dict[int, dict[str, Any]],
    products_by_template: dict[int, list[dict[str, Any]]],
) -> tuple[
    dict[str, SelectedBom],
    dict[int, dict[str, Any]],
]:
    rows = client.search_read(
        model="mrp.bom",
        domain=[
            ["active", "=", True],
            ["sequence", "=", 0],
        ],
        fields=[
            "code",
            "type",
            "sequence",
            "product_id",
            "product_tmpl_id",
            "product_qty",
            "bom_line_ids",
            "write_date",
        ],
        limit=MAX_RECORDS,
    )

    grouped: dict[
        str,
        list[dict[str, Any]],
    ] = defaultdict(list)

    by_id = {
        int(row["id"]): row
        for row in rows
    }

    for row in rows:
        parent_sku = resolve_bom_parent_sku(
            row,
            products_by_id,
            products_by_template,
        )

        if parent_sku:
            grouped[parent_sku].append(
                row
            )

    selected: dict[str, SelectedBom] = {}

    for parent_sku, candidates in grouped.items():
        chosen = max(
            candidates,
            key=lambda row: (
                parse_write_date(
                    row.get("write_date")
                ),
                int(row["id"]),
            ),
        )

        selected[parent_sku] = SelectedBom(
            bom_id=int(chosen["id"]),
            parent_sku=parent_sku,
            reference=str(
                chosen.get("code")
                or ""
            ).strip(),
            bom_type=str(
                chosen.get("type")
                or ""
            ).strip(),
            sequence=int(
                chosen.get("sequence")
                or 0
            ),
            product_qty=to_float(
                chosen.get("product_qty")
            ),
            write_date=str(
                chosen.get("write_date")
                or ""
            ),
            duplicate_sequence0_count=len(
                candidates
            ),
        )

    return selected, by_id


def load_odoo_components(
    client: OdooClient,
    *,
    selected_boms: dict[str, SelectedBom],
    products_by_id: dict[int, dict[str, Any]],
) -> dict[str, dict[str, float]]:
    bom_ids = sorted(
        bom.bom_id
        for bom in selected_boms.values()
    )

    if not bom_ids:
        return {}

    lines = client.search_read(
        model="mrp.bom.line",
        domain=[
            ["bom_id", "in", bom_ids],
        ],
        fields=[
            "bom_id",
            "product_id",
            "product_qty",
        ],
        limit=MAX_RECORDS,
    )

    parent_by_bom_id = {
        bom.bom_id: parent_sku
        for parent_sku, bom in (
            selected_boms.items()
        )
    }

    result: dict[
        str,
        dict[str, float],
    ] = defaultdict(
        lambda: defaultdict(float)
    )

    for line in lines:
        bom_id = many2one_id(
            line.get("bom_id")
        )

        if bom_id is None:
            continue

        parent_sku = parent_by_bom_id.get(
            bom_id
        )

        if not parent_sku:
            continue

        product_id = many2one_id(
            line.get("product_id")
        )

        if product_id is None:
            continue

        product = products_by_id.get(
            product_id,
            {},
        )

        component_sku = canon(
            product.get("default_code")
        )

        if not component_sku:
            continue

        base_qty = selected_boms[
            parent_sku
        ].product_qty

        if base_qty <= 0:
            continue

        normalized_qty = (
            to_float(
                line.get("product_qty")
            )
            / base_qty
        )

        result[parent_sku][
            component_sku
        ] += normalized_qty

    return {
        parent_sku: dict(components)
        for parent_sku, components
        in result.items()
    }


def load_catalog_components() -> tuple[
    dict[str, dict[str, float]],
    str,
    str,
]:
    dataset = (
        ValidatedDatasetRepository()
        .load_latest("production")
    )

    result: dict[
        str,
        dict[str, float],
    ] = {}

    for product in dataset.products:
        parent_sku = canon(
            product.sku
        )

        components: dict[
            str,
            float,
        ] = defaultdict(float)

        for component in product.components:
            component_sku = canon(
                component.get("sku")
            )

            if not component_sku:
                continue

            components[
                component_sku
            ] += to_float(
                component.get("quantity")
            )

        result[parent_sku] = dict(
            components
        )

    return (
        result,
        dataset.dataset_id,
        dataset.batch_reference,
    )


def compare_boms(
    *,
    catalog_components: dict[str, dict[str, float]],
    selected_boms: dict[str, SelectedBom],
    odoo_components: dict[str, dict[str, float]],
) -> list[DifferenceRow]:
    rows: list[DifferenceRow] = []

    for parent_sku in sorted(
        catalog_components
    ):
        catalog = catalog_components[
            parent_sku
        ]

        selected_bom = selected_boms.get(
            parent_sku
        )

        if selected_bom is None:
            rows.append(
                DifferenceRow(
                    parent_sku=parent_sku,
                    bom_id=None,
                    bom_reference="",
                    status="MISSING ACTIVE SEQUENCE 0 BOM",
                    component_sku="",
                    catalog_qty=0.0,
                    odoo_qty=0.0,
                    difference=0.0,
                    duplicate_sequence0_count=0,
                )
            )
            continue

        odoo = odoo_components.get(
            parent_sku,
            {},
        )

        all_skus = sorted(
            set(catalog)
            | set(odoo)
        )

        parent_has_difference = False

        for component_sku in all_skus:
            catalog_qty = catalog.get(
                component_sku,
                0.0,
            )
            odoo_qty = odoo.get(
                component_sku,
                0.0,
            )
            difference = (
                odoo_qty - catalog_qty
            )

            if (
                abs(catalog_qty)
                <= QTY_TOLERANCE
                and abs(odoo_qty)
                > QTY_TOLERANCE
            ):
                status = "EXTRA IN ODOO BOM"
            elif (
                abs(catalog_qty)
                > QTY_TOLERANCE
                and abs(odoo_qty)
                <= QTY_TOLERANCE
            ):
                status = "MISSING IN ODOO BOM"
            elif abs(difference) > QTY_TOLERANCE:
                status = "QTY MISMATCH"
            else:
                continue

            parent_has_difference = True

            rows.append(
                DifferenceRow(
                    parent_sku=parent_sku,
                    bom_id=selected_bom.bom_id,
                    bom_reference=(
                        selected_bom.reference
                    ),
                    status=status,
                    component_sku=component_sku,
                    catalog_qty=catalog_qty,
                    odoo_qty=odoo_qty,
                    difference=difference,
                    duplicate_sequence0_count=(
                        selected_bom
                        .duplicate_sequence0_count
                    ),
                )
            )

        if (
            not parent_has_difference
            and selected_bom
            .duplicate_sequence0_count > 1
        ):
            rows.append(
                DifferenceRow(
                    parent_sku=parent_sku,
                    bom_id=selected_bom.bom_id,
                    bom_reference=(
                        selected_bom.reference
                    ),
                    status="DUPLICATE ACTIVE SEQUENCE 0 BOM",
                    component_sku="",
                    catalog_qty=0.0,
                    odoo_qty=0.0,
                    difference=0.0,
                    duplicate_sequence0_count=(
                        selected_bom
                        .duplicate_sequence0_count
                    ),
                )
            )

    return rows


def build_parent_summary(
    rows: list[DifferenceRow],
) -> list[dict[str, Any]]:
    grouped: dict[
        str,
        list[DifferenceRow],
    ] = defaultdict(list)

    for row in rows:
        grouped[
            row.parent_sku
        ].append(
            row
        )

    result = []

    for parent_sku, parent_rows in sorted(
        grouped.items()
    ):
        statuses = Counter(
            row.status
            for row in parent_rows
        )

        first = parent_rows[0]

        result.append(
            {
                "parent_sku": parent_sku,
                "bom_id": first.bom_id,
                "bom_reference": (
                    first.bom_reference
                ),
                "missing": statuses[
                    "MISSING IN ODOO BOM"
                ],
                "extra": statuses[
                    "EXTRA IN ODOO BOM"
                ],
                "qty_mismatch": statuses[
                    "QTY MISMATCH"
                ],
                "missing_active_bom": statuses[
                    "MISSING ACTIVE SEQUENCE 0 BOM"
                ],
                "duplicate_sequence0": max(
                    row.duplicate_sequence0_count
                    for row in parent_rows
                ),
                "status": (
                    "CORRECTION REQUIRED"
                ),
            }
        )

    return result


def style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    fill = PatternFill(
        "solid",
        fgColor="1F4E78",
    )

    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(
            color="FFFFFF",
            bold=True,
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for column in range(
        1,
        ws.max_column + 1,
    ):
        width = max(
            len(
                str(
                    ws.cell(
                        row=row,
                        column=column,
                    ).value
                    or ""
                )
            )
            for row in range(
                1,
                min(ws.max_row, 500) + 1,
            )
        )

        ws.column_dimensions[
            get_column_letter(column)
        ].width = min(
            max(width + 2, 12),
            60,
        )


def write_report(
    *,
    differences: list[DifferenceRow],
    parent_summary: list[dict[str, Any]],
    dataset_id: str,
    batch_reference: str,
    output_dir: Path,
) -> Path:
    output_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    timestamp = datetime.now().strftime(
        "%Y%m%d_%H%M%S"
    )

    output_path = (
        output_dir
        / (
            "BOM_Sync_Correction_Report_"
            f"{timestamp}.xlsx"
        )
    )

    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "SUMMARY"
    summary_ws.append(
        ["Metric", "Value"]
    )
    summary_ws.append(
        ["Dataset ID", dataset_id]
    )
    summary_ws.append(
        ["Dataset Batch", batch_reference]
    )
    summary_ws.append(
        [
            "BOM requiring correction",
            len(parent_summary),
        ]
    )
    summary_ws.append(
        [
            "Difference rows",
            len(differences),
        ]
    )

    parent_ws = wb.create_sheet(
        "BOMS_TO_CORRECT"
    )
    parent_ws.append(
        [
            "Parent SKU",
            "Odoo BOM ID",
            "Odoo BOM Reference",
            "Missing Components",
            "Extra Components",
            "Qty Mismatches",
            "Missing Active Sequence 0 BOM",
            "Active Sequence 0 BOM Count",
            "Status",
        ]
    )

    for row in parent_summary:
        parent_ws.append(
            [
                row["parent_sku"],
                row["bom_id"],
                row["bom_reference"],
                row["missing"],
                row["extra"],
                row["qty_mismatch"],
                row["missing_active_bom"],
                row["duplicate_sequence0"],
                row["status"],
            ]
        )

    details_ws = wb.create_sheet(
        "CORRECTION_LINES"
    )
    details_ws.append(
        [
            "Parent SKU",
            "Odoo BOM ID",
            "Odoo BOM Reference",
            "Status",
            "Component SKU",
            "Catalog Qty",
            "Odoo Qty",
            "Odoo - Catalog",
            "Active Sequence 0 BOM Count",
        ]
    )

    for row in differences:
        details_ws.append(
            [
                row.parent_sku,
                row.bom_id,
                row.bom_reference,
                row.status,
                row.component_sku,
                row.catalog_qty,
                row.odoo_qty,
                row.difference,
                row.duplicate_sequence0_count,
            ]
        )

    for ws in (
        summary_ws,
        parent_ws,
        details_ws,
    ):
        style_sheet(ws)

    wb.save(output_path)

    return output_path.resolve()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Parengia koreguotinų aktyvių Sequence 0 "
            "Odoo BOM sąrašą pagal Product Catalog."
        )
    )

    parser.add_argument(
        "--output-dir",
        type=Path,
        default=(
            Path(__file__).resolve().parents[1]
            / "output"
        ),
    )

    return parser.parse_args()


def main() -> None:
    args = parse_args()

    client = OdooClient(
        OdooConfig.from_env()
    )

    try:
        client.connect()

        (
            catalog_components,
            dataset_id,
            batch_reference,
        ) = load_catalog_components()

        (
            products_by_id,
            products_by_template,
        ) = load_products(
            client
        )

        (
            selected_boms,
            _,
        ) = choose_active_sequence0_boms(
            client,
            products_by_id,
            products_by_template,
        )

        odoo_components = load_odoo_components(
            client,
            selected_boms=selected_boms,
            products_by_id=products_by_id,
        )

        differences = compare_boms(
            catalog_components=catalog_components,
            selected_boms=selected_boms,
            odoo_components=odoo_components,
        )

        parent_summary = build_parent_summary(
            differences
        )

        output_path = write_report(
            differences=differences,
            parent_summary=parent_summary,
            dataset_id=dataset_id,
            batch_reference=batch_reference,
            output_dir=args.output_dir,
        )

        statuses = Counter(
            row.status
            for row in differences
        )

        print()
        print("=" * 80)
        print("BOM SYNC CORRECTION REPORT")
        print("=" * 80)
        print(
            "Koreguotinų BOM:",
            len(parent_summary),
        )
        print(
            "Neatitikimų eilučių:",
            len(differences),
        )

        for status, count in sorted(
            statuses.items()
        ):
            print(
                f"{status:<36} {count}"
            )

        print()
        print("Ataskaita:", output_path)

    except OdooConnectionError as exc:
        print(
            f"\nOdoo klaida: {exc}"
        )
        raise SystemExit(1) from exc
    except Exception as exc:
        print(
            f"\nNenumatyta klaida: {exc}"
        )
        raise SystemExit(1) from exc


if __name__ == "__main__":
    main()
