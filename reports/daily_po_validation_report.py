from __future__ import annotations

import argparse
from collections import Counter
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.bom_engine import BomEngine
from core.bom_repository import BomRepository
from core.bom_selector import BomSelector
from core.config import OdooConfig
from core.dataset_execution_engine import DatasetExecutionEngine
from core.odoo_client import OdooClient, OdooConnectionError
from core.odoo_helpers import many2one_id, many2one_name
from validators.furnix_po_validator import (
    FurnixPoValidationResult,
    FurnixPoValidator,
)


FURNIX_VENDOR_NAME = "FURNIX, UAB"
LOCAL_TIMEZONE = ZoneInfo("Europe/Vilnius")

PENDING_STATES = {"draft", "sent", "to approve"}
APPROVED_STATES = {"purchase", "done"}
READY_STATUSES = {"PASS", "PASS WITH FALLBACK"}


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def parse_iso_date(value: str | None) -> date:
    if not value:
        return datetime.now(LOCAL_TIMEZONE).date()

    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        raise ValueError(
            f"Neteisinga data {value!r}. Naudok YYYY-MM-DD."
        ) from exc


def local_day_to_utc_strings(
    selected_date: date,
) -> tuple[str, str]:
    local_start = datetime.combine(
        selected_date,
        time.min,
        tzinfo=LOCAL_TIMEZONE,
    )
    local_end = local_start + timedelta(days=1)

    utc_start = local_start.astimezone(
        timezone.utc
    ).replace(tzinfo=None)

    utc_end = local_end.astimezone(
        timezone.utc
    ).replace(tzinfo=None)

    return (
        utc_start.strftime("%Y-%m-%d %H:%M:%S"),
        utc_end.strftime("%Y-%m-%d %H:%M:%S"),
    )


def build_validator(
    client: OdooClient,
) -> FurnixPoValidator:
    bom_repository = BomRepository(client)
    bom_selector = BomSelector(bom_repository)
    odoo_bom_engine = BomEngine(
        bom_repository,
        bom_selector,
    )

    execution_engine = DatasetExecutionEngine(
        client,
        odoo_bom_engine,
        environment="production",
    )

    return FurnixPoValidator(
        client,
        execution_engine,
    )


def load_purchase_orders(
    client: OdooClient,
    *,
    mode: str,
    selected_date: date,
    so_number: str = "",
    po_number: str = "",
) -> list[dict[str, Any]]:
    normalized_mode = normalize_text(mode).lower()

    domain: list[list[Any]] = [
        ["state", "!=", "cancel"],
        ["sale_primary_id", "!=", False],
    ]

    if normalized_mode == "pending":
        domain.append(
            ["state", "in", sorted(PENDING_STATES)]
        )

    elif normalized_mode == "today-created":
        utc_start, utc_end = local_day_to_utc_strings(
            selected_date
        )
        domain.extend(
            [
                ["create_date", ">=", utc_start],
                ["create_date", "<", utc_end],
            ]
        )

    elif normalized_mode == "today-approved":
        utc_start, utc_end = local_day_to_utc_strings(
            selected_date
        )
        domain.extend(
            [
                ["state", "in", sorted(APPROVED_STATES)],
                ["date_approve", ">=", utc_start],
                ["date_approve", "<", utc_end],
            ]
        )

    elif normalized_mode == "so":
        normalized_so = normalize_text(
            so_number
        )
        if not normalized_so:
            raise ValueError(
                "--mode so reikalauja --so-number."
            )

        sale_orders = client.search_read(
            model="sale.order",
            domain=[
                ["name", "=", normalized_so],
            ],
            fields=[
                "name",
            ],
            limit=2,
        )

        if not sale_orders:
            return []

        if len(sale_orders) > 1:
            raise RuntimeError(
                f"Rasti keli SO numeriu {normalized_so}."
            )

        domain.append(
            ["sale_primary_id", "=", int(sale_orders[0]["id"])]
        )

    elif normalized_mode == "po":
        normalized_po = normalize_text(
            po_number
        )
        if not normalized_po:
            raise ValueError(
                "--mode po reikalauja --po-number."
            )

        domain.append(
            ["name", "=", normalized_po]
        )

    else:
        raise ValueError(
            f"Neatpažintas režimas: {mode!r}."
        )

    purchase_orders = client.search_read(
        model="purchase.order",
        domain=domain,
        fields=[
            "name",
            "state",
            "partner_id",
            "sale_primary_id",
            "date_order",
            "date_approve",
            "create_date",
        ],
        order="id asc",
    )

    furnix_orders = [
        po
        for po in purchase_orders
        if many2one_name(
            po.get("partner_id")
        ).strip().upper() == FURNIX_VENDOR_NAME
    ]

    return furnix_orders


def purchase_orders_to_sale_orders(
    client: OdooClient,
    purchase_orders: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    sale_order_ids = sorted(
        {
            sale_order_id
            for po in purchase_orders
            if (
                sale_order_id := many2one_id(
                    po.get("sale_primary_id")
                )
            )
            is not None
        }
    )

    if not sale_order_ids:
        return []

    sale_orders = client.read(
        model="sale.order",
        record_ids=sale_order_ids,
        fields=[
            "name",
            "date_order",
            "partner_id",
        ],
    )

    sale_orders_by_id = {
        int(sale_order["id"]): sale_order
        for sale_order in sale_orders
    }

    rows: list[dict[str, Any]] = []

    for po in purchase_orders:
        sale_order_id = many2one_id(
            po.get("sale_primary_id")
        )

        sale_order = sale_orders_by_id.get(
            sale_order_id or -1,
            {},
        )

        so_number = normalize_text(
            sale_order.get("name")
            or many2one_name(
                po.get("sale_primary_id")
            )
        )

        if not so_number:
            continue

        rows.append(
            {
                "so_number": so_number,
                "po_number": normalize_text(
                    po.get("name")
                ),
                "po_state": normalize_text(
                    po.get("state")
                ),
                "po_created": po.get(
                    "create_date"
                ),
                "po_approved": po.get(
                    "date_approve"
                ),
                "customer": many2one_name(
                    sale_order.get("partner_id")
                ),
            }
        )

    unique_by_so: dict[str, dict[str, Any]] = {}

    for row in rows:
        unique_by_so.setdefault(
            row["so_number"],
            row,
        )

    return sorted(
        unique_by_so.values(),
        key=lambda row: (
            row["so_number"],
            row["po_number"],
        ),
    )


def validate_orders(
    validator: FurnixPoValidator,
    sale_orders: list[dict[str, Any]],
    *,
    verbose: bool,
) -> list[tuple[dict[str, Any], FurnixPoValidationResult]]:
    results = []
    total = len(sale_orders)

    for index, sale_order in enumerate(
        sale_orders,
        start=1,
    ):
        so_number = sale_order["so_number"]

        if verbose:
            print(
                f"Tikrinama {index}/{total}: "
                f"{so_number}"
            )

        try:
            result = validator.validate(
                so_number
            )
        except Exception as exc:
            result = FurnixPoValidationResult(
                so_number=so_number,
                po_number=sale_order.get(
                    "po_number",
                    "",
                ),
                po_state=sale_order.get(
                    "po_state",
                    "",
                ),
                status="VALIDATOR ERROR",
                error=str(exc),
            )

        results.append(
            (sale_order, result)
        )

    return results


def style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_fill = PatternFill(
        "solid",
        fgColor="1F4E78",
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(
            color="FFFFFF",
            bold=True,
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for column in range(
        1,
        ws.max_column + 1,
    ):
        values = [
            len(
                str(
                    ws.cell(
                        row=row,
                        column=column,
                    ).value
                    or ""
                )
            )
            for row in range(
                1,
                min(ws.max_row, 500) + 1,
            )
        ]

        ws.column_dimensions[
            get_column_letter(column)
        ].width = min(
            max(values, default=12) + 2,
            55,
        )


def write_report(
    *,
    mode: str,
    selected_date: date,
    results: list[
        tuple[
            dict[str, Any],
            FurnixPoValidationResult,
        ]
    ],
    output_dir: Path,
) -> Path:
    output_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    timestamp = datetime.now().strftime(
        "%Y%m%d_%H%M%S"
    )

    output_path = (
        output_dir
        / (
            f"Furnix_PO_Validation_"
            f"{mode}_{selected_date:%Y%m%d}_"
            f"{timestamp}.xlsx"
        )
    )

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "SUMMARY"

    summary_ws.append(
        [
            "SO",
            "PO",
            "PO State",
            "PO Created",
            "PO Approved",
            "Customer",
            "Status",
            "Ready to Approve",
            "Fallback Count",
            "PASS",
            "MISSING",
            "EXTRA",
            "QTY MISMATCH",
            "STICKER ERROR",
            "Dataset ID",
            "Dataset Batch",
            "Warnings",
            "Error",
        ]
    )

    detail_ws = wb.create_sheet(
        "DETAILS"
    )
    detail_ws.append(
        [
            "SO",
            "PO",
            "Validation Status",
            "SKU",
            "Row Status",
            "Sticker Status",
            "Required Qty",
            "PO Qty",
            "Difference",
            "Sticker Error",
        ]
    )

    fallback_ws = wb.create_sheet(
        "FALLBACKS"
    )
    fallback_ws.append(
        [
            "SO",
            "PO",
            "Root SKU",
            "Group",
            "Reason",
            "Source",
        ]
    )

    for source_row, result in results:
        is_ready = (
            result.status in READY_STATUSES
        )

        summary_ws.append(
            [
                result.so_number,
                result.po_number
                or source_row.get(
                    "po_number",
                    "",
                ),
                result.po_state
                or source_row.get(
                    "po_state",
                    "",
                ),
                source_row.get(
                    "po_created",
                    "",
                ),
                source_row.get(
                    "po_approved",
                    "",
                ),
                source_row.get(
                    "customer",
                    "",
                ),
                result.status,
                "YES" if is_ready else "NO",
                result.fallback_count,
                result.pass_count,
                result.missing_count,
                result.extra_count,
                result.qty_mismatch_count,
                result.sticker_error_count,
                result.dataset_id,
                result.batch_reference,
                "; ".join(
                    result.warnings
                ),
                result.error,
            ]
        )

        for row in result.rows:
            detail_ws.append(
                [
                    result.so_number,
                    result.po_number,
                    result.status,
                    row.sku,
                    row.status,
                    row.sticker_status,
                    row.required_qty,
                    row.po_qty,
                    row.difference,
                    row.sticker_error,
                ]
            )

        for fallback in result.fallbacks:
            fallback_ws.append(
                [
                    result.so_number,
                    result.po_number,
                    fallback.get(
                        "root_sku",
                        "",
                    ),
                    fallback.get(
                        "group_name",
                        "",
                    ),
                    fallback.get(
                        "reason",
                        "",
                    ),
                    fallback.get(
                        "source",
                        "",
                    ),
                ]
            )

    style_sheet(summary_ws)
    style_sheet(detail_ws)
    style_sheet(fallback_ws)

    wb.save(output_path)
    return output_path.resolve()


def print_summary(
    *,
    mode: str,
    selected_date: date,
    results: list[
        tuple[
            dict[str, Any],
            FurnixPoValidationResult,
        ]
    ],
    output_path: Path,
) -> None:
    statuses = Counter(
        result.status
        for _, result in results
    )

    ready = sum(
        result.status in READY_STATUSES
        for _, result in results
    )
    blocked = len(results) - ready
    fallback_orders = sum(
        result.fallback_count > 0
        for _, result in results
    )

    print()
    print("=" * 80)
    print("FURNIX PO VALIDATION SUMMARY")
    print("=" * 80)
    print("Režimas:", mode)
    print("Data:", selected_date)
    print("Patikrinta:", len(results))
    print("Galima tvirtinti:", ready)
    print("Reikia taisyti:", blocked)
    print("SO su Odoo fallback:", fallback_orders)

    if statuses:
        print()
        for status, count in sorted(
            statuses.items()
        ):
            print(
                f"{status:<24} {count}"
            )

    print()
    print("GALIMA TVIRTINTI:")

    ready_rows = [
        result
        for _, result in results
        if result.status in READY_STATUSES
    ]

    if ready_rows:
        for result in ready_rows:
            suffix = (
                f" | fallback={result.fallback_count}"
                if result.fallback_count
                else ""
            )

            print(
                f"- {result.so_number} / "
                f"{result.po_number}: "
                f"{result.status}{suffix}"
            )
    else:
        print("- nėra")

    print()
    print("REIKIA TAISYTI:")

    blocked_rows = [
        result
        for _, result in results
        if result.status not in READY_STATUSES
    ]

    if blocked_rows:
        for result in blocked_rows:
            detail = (
                result.error
                or (
                    f"missing={result.missing_count}, "
                    f"extra={result.extra_count}, "
                    f"qty={result.qty_mismatch_count}, "
                    f"sticker={result.sticker_error_count}"
                )
            )

            print(
                f"- {result.so_number} / "
                f"{result.po_number or '-'}: "
                f"{result.status} | {detail}"
            )
    else:
        print("- nėra")

    print()
    print("Ataskaita:", output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Aktuali Furnix PO patikra per Product Catalog."
        )
    )

    parser.add_argument(
        "--mode",
        choices=[
            "pending",
            "today-created",
            "today-approved",
            "so",
            "po",
        ],
        default="pending",
    )

    parser.add_argument(
        "--date",
        help=(
            "Data YYYY-MM-DD. Nenurodžius naudojama šiandienos "
            "Europe/Vilnius data."
        ),
    )

    parser.add_argument(
        "--so-number",
        default="",
    )

    parser.add_argument(
        "--po-number",
        default="",
    )

    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Rodyti kiekvieną tikrinamą SO.",
    )

    parser.add_argument(
        "--output-dir",
        type=Path,
        default=(
            Path(__file__).resolve().parents[1]
            / "output"
        ),
    )

    return parser.parse_args()


def main() -> None:
    args = parse_args()
    selected_date = parse_iso_date(
        args.date
    )

    client = OdooClient(
        OdooConfig.from_env()
    )

    try:
        client.connect()

        purchase_orders = load_purchase_orders(
            client,
            mode=args.mode,
            selected_date=selected_date,
            so_number=args.so_number,
            po_number=args.po_number,
        )

        sale_orders = purchase_orders_to_sale_orders(
            client,
            purchase_orders,
        )

        if not sale_orders:
            print(
                "Tikrintinų Furnix PO nerasta."
            )
            return

        validator = build_validator(
            client
        )

        results = validate_orders(
            validator,
            sale_orders,
            verbose=args.verbose,
        )

        output_path = write_report(
            mode=args.mode,
            selected_date=selected_date,
            results=results,
            output_dir=args.output_dir,
        )

        print_summary(
            mode=args.mode,
            selected_date=selected_date,
            results=results,
            output_path=output_path,
        )

    except OdooConnectionError as exc:
        print(
            f"\nOdoo klaida: {exc}"
        )
        raise SystemExit(1) from exc
    except Exception as exc:
        print(
            f"\nNenumatyta klaida: {exc}"
        )
        raise SystemExit(1) from exc


if __name__ == "__main__":
    main()