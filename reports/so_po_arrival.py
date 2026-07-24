from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Font, PatternFill

from core.report_base import BaseReport


class SoPoArrivalReport(BaseReport):
    name = "SO → PO gavimo datų ataskaita"
    description = (
        "Aktyvūs pardavimo užsakymai, susiję aktyvūs pirkimo užsakymai "
        "ir jų planuojamos bei faktinės gavimo datos."
    )

    def run(self) -> None:
        print("\n====================================")
        print(self.name)
        print("====================================")
        print("\nNuskaitomi duomenys iš Odoo...")

        dataframe = self.build_dataframe()

        if dataframe.empty:
            print("\nAtaskaitai tinkamų duomenų nerasta.")
            return

        self.print_preview(dataframe)

        output_path = self.export_to_excel(dataframe)

        print("\nAtaskaita sukurta:")
        print(output_path)

    def build_dataframe(self) -> pd.DataFrame:
        sale_orders = self.load_sale_orders()
        rows: list[dict[str, Any]] = []

        for index, sale_order in enumerate(sale_orders, start=1):
            print(
                f"\rApdorojamas SO {index}/{len(sale_orders)}",
                end="",
                flush=True,
            )

            purchase_orders = self.load_purchase_orders(
                sale_order_id=sale_order["id"]
            )

            if not purchase_orders:
                rows.append(
                    self.build_row(
                        sale_order=sale_order,
                        purchase_order=None,
                        receipt=None,
                    )
                )
                continue

            for purchase_order in purchase_orders:
                receipts = self.load_incoming_receipts(
                    purchase_order.get("picking_ids", [])
                )

                if not receipts:
                    rows.append(
                        self.build_row(
                            sale_order=sale_order,
                            purchase_order=purchase_order,
                            receipt=None,
                        )
                    )
                    continue

                for receipt in receipts:
                    rows.append(
                        self.build_row(
                            sale_order=sale_order,
                            purchase_order=purchase_order,
                            receipt=receipt,
                        )
                    )

        print()

        dataframe = pd.DataFrame(rows)

        if dataframe.empty:
            return dataframe

        dataframe = dataframe.sort_values(
            by=[
                "Risk Priority",
                "Days Until Commitment",
                "PO Expected Arrival",
                "SO",
                "PO",
                "Receipt",
            ],
            ascending=[
                True,
                True,
                True,
                True,
                True,
                True,
            ],
            na_position="last",
        ).reset_index(drop=True)

        dataframe = dataframe.drop(
            columns=["Risk Priority"]
        )

        return dataframe

    def load_sale_orders(self) -> list[dict[str, Any]]:
        domain = [
            ("state", "=", "sale"),
            ("invoice_status", "!=", "invoiced"),
        ]

        fields = [
            "id",
            "name",
            "partner_id",
            "date_order",
            "commitment_date",
            "invoice_status",
        ]

        return self.client.search_read(
            model="sale.order",
            domain=domain,
            fields=fields,
            order="date_order desc",
        )

    def load_purchase_orders(
        self,
        sale_order_id: int,
    ) -> list[dict[str, Any]]:
        domain = [
            ("sale_primary_id", "=", sale_order_id),
            ("state", "!=", "cancel"),
        ]

        fields = [
            "id",
            "name",
            "partner_id",
            "state",
            "date_planned",
            "picking_ids",
        ]

        return self.client.search_read(
            model="purchase.order",
            domain=domain,
            fields=fields,
            order="date_planned asc",
        )

    def load_incoming_receipts(
        self,
        picking_ids: list[int],
    ) -> list[dict[str, Any]]:
        if not picking_ids:
            return []

        receipts = self.client.read(
            model="stock.picking",
            record_ids=picking_ids,
            fields=[
                "name",
                "state",
                "picking_type_code",
                "scheduled_date",
                "date_done",
                "origin",
            ],
        )

        return [
            receipt
            for receipt in receipts
            if receipt.get("picking_type_code") == "incoming"
            and receipt.get("state") != "cancel"
        ]

    def build_row(
        self,
        sale_order: dict[str, Any],
        purchase_order: dict[str, Any] | None,
        receipt: dict[str, Any] | None,
    ) -> dict[str, Any]:
        today = pd.Timestamp.now().normalize()

        so_date = self.to_datetime(
            sale_order.get("date_order")
        )
        commitment_date = self.to_datetime(
            sale_order.get("commitment_date")
        )

        so_age_days = self.calculate_date_difference(
            later_date=today,
            earlier_date=so_date,
        )

        days_until_commitment = self.calculate_date_difference(
            later_date=commitment_date,
            earlier_date=today,
        )

        base_row = {
            "SO": sale_order.get("name", ""),
            "Customer": self.get_many2one_name(
                sale_order.get("partner_id")
            ),
            "SO Date": so_date,
            "SO Age (days)": so_age_days,
            "Commitment Date": commitment_date,
            "Days Until Commitment": days_until_commitment,
            "Invoice Status": sale_order.get(
                "invoice_status",
                "",
            ),
        }

        if purchase_order is None:
            return {
                **base_row,
                "PO": "",
                "Vendor": "",
                "PO State": "No Active Purchase Order",
                "PO Expected Arrival": pd.NaT,
                "Receipt": "",
                "Receipt State": "",
                "Scheduled Arrival": pd.NaT,
                "Actual Arrival": pd.NaT,
                "Supplier Delay (days)": pd.NA,
                "Risk Status": "Missing PO",
                "Risk Priority": 1,
            }

        expected_arrival = self.to_datetime(
            purchase_order.get("date_planned")
        )

        po_row = {
            **base_row,
            "PO": purchase_order.get("name", ""),
            "Vendor": self.get_many2one_name(
                purchase_order.get("partner_id")
            ),
            "PO State": purchase_order.get("state", ""),
            "PO Expected Arrival": expected_arrival,
        }

        if receipt is None:
            delay_days = self.calculate_open_delay(
                expected_arrival=expected_arrival,
                today=today,
            )

            risk_status, risk_priority = self.calculate_risk_status(
                days_until_commitment=days_until_commitment,
                supplier_delay_days=delay_days,
                receipt_state="",
                has_purchase_order=True,
            )

            return {
                **po_row,
                "Receipt": "",
                "Receipt State": "No Incoming Receipt",
                "Scheduled Arrival": pd.NaT,
                "Actual Arrival": pd.NaT,
                "Supplier Delay (days)": delay_days,
                "Risk Status": risk_status,
                "Risk Priority": risk_priority,
            }

        receipt_state = str(
            receipt.get("state", "")
        )
        actual_arrival = self.to_datetime(
            receipt.get("date_done")
        )

        supplier_delay_days = self.calculate_supplier_delay(
            expected_arrival=expected_arrival,
            actual_arrival=actual_arrival,
            receipt_state=receipt_state,
            today=today,
        )

        risk_status, risk_priority = self.calculate_risk_status(
            days_until_commitment=days_until_commitment,
            supplier_delay_days=supplier_delay_days,
            receipt_state=receipt_state,
            has_purchase_order=True,
        )

        return {
            **po_row,
            "Receipt": receipt.get("name", ""),
            "Receipt State": receipt_state,
            "Scheduled Arrival": self.to_datetime(
                receipt.get("scheduled_date")
            ),
            "Actual Arrival": actual_arrival,
            "Supplier Delay (days)": supplier_delay_days,
            "Risk Status": risk_status,
            "Risk Priority": risk_priority,
        }

    @staticmethod
    def calculate_date_difference(
        later_date: Any,
        earlier_date: Any,
    ) -> Any:
        if pd.isna(later_date) or pd.isna(earlier_date):
            return pd.NA

        later_day = pd.Timestamp(later_date).normalize()
        earlier_day = pd.Timestamp(earlier_date).normalize()

        return int(
            (later_day - earlier_day).days
        )

    @staticmethod
    def calculate_open_delay(
        expected_arrival: Any,
        today: pd.Timestamp,
    ) -> Any:
        if pd.isna(expected_arrival):
            return pd.NA

        expected_day = pd.Timestamp(
            expected_arrival
        ).normalize()

        return max(
            int((today - expected_day).days),
            0,
        )

    @staticmethod
    def calculate_supplier_delay(
        expected_arrival: Any,
        actual_arrival: Any,
        receipt_state: str,
        today: pd.Timestamp,
    ) -> Any:
        if pd.isna(expected_arrival):
            return pd.NA

        expected_day = pd.Timestamp(
            expected_arrival
        ).normalize()

        if (
            receipt_state == "done"
            and not pd.isna(actual_arrival)
        ):
            actual_day = pd.Timestamp(
                actual_arrival
            ).normalize()

            return max(
                int((actual_day - expected_day).days),
                0,
            )

        return max(
            int((today - expected_day).days),
            0,
        )

    @staticmethod
    def calculate_risk_status(
        days_until_commitment: Any,
        supplier_delay_days: Any,
        receipt_state: str,
        has_purchase_order: bool,
    ) -> tuple[str, int]:
        if not has_purchase_order:
            return "Missing PO", 1

        if receipt_state == "done":
            if (
                not pd.isna(supplier_delay_days)
                and int(supplier_delay_days) > 0
            ):
                return "Received Late", 5

            return "Received", 6

        if (
            not pd.isna(days_until_commitment)
            and int(days_until_commitment) < 0
        ):
            return "Customer Late", 2

        if (
            not pd.isna(supplier_delay_days)
            and int(supplier_delay_days) > 0
        ):
            return "Supplier Late", 3

        return "Waiting", 4

    @staticmethod
    def get_many2one_name(value: Any) -> str:
        if (
            isinstance(value, (list, tuple))
            and len(value) > 1
        ):
            return str(value[1])

        return ""

    @staticmethod
    def to_datetime(value: Any) -> Any:
        if not value:
            return pd.NaT

        return pd.to_datetime(
            value,
            errors="coerce",
        )

    @staticmethod
    def print_preview(
        dataframe: pd.DataFrame,
        row_limit: int = 30,
    ) -> None:
        print(f"\nAtaskaitos eilučių: {len(dataframe)}")
        print("\nRizikų suvestinė:")

        status_counts = dataframe[
            "Risk Status"
        ].value_counts(dropna=False)

        for status, count in status_counts.items():
            print(f"- {status}: {count}")

        print("\nPirmos ataskaitos eilutės:\n")

        with pd.option_context(
            "display.max_rows",
            row_limit,
            "display.max_columns",
            None,
            "display.width",
            320,
        ):
            print(
                dataframe.head(row_limit).to_string(
                    index=False
                )
            )

    @staticmethod
    def export_to_excel(
        dataframe: pd.DataFrame,
    ) -> Path:
        output_directory = Path("output")
        output_directory.mkdir(
            parents=True,
            exist_ok=True,
        )

        output_path = (
            output_directory
            / "SO_PO_Arrival_Report.xlsx"
        )

        with pd.ExcelWriter(
            output_path,
            engine="openpyxl",
            datetime_format="yyyy-mm-dd hh:mm:ss",
        ) as writer:
            dataframe.to_excel(
                writer,
                index=False,
                sheet_name="SO PO Arrival",
            )

            worksheet = writer.sheets["SO PO Arrival"]

            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions

            column_widths = {
                "A": 22,
                "B": 35,
                "C": 22,
                "D": 16,
                "E": 22,
                "F": 23,
                "G": 18,
                "H": 16,
                "I": 30,
                "J": 24,
                "K": 22,
                "L": 22,
                "M": 22,
                "N": 22,
                "O": 22,
                "P": 24,
            }

            for column, width in column_widths.items():
                worksheet.column_dimensions[column].width = width

            for cell in worksheet[1]:
                cell.font = Font(bold=True)

            status_column = None

            for cell in worksheet[1]:
                if cell.value == "Risk Status":
                    status_column = cell.column
                    break

            fills = {
                "Missing PO": PatternFill(
                    fill_type="solid",
                    fgColor="F4CCCC",
                ),
                "Customer Late": PatternFill(
                    fill_type="solid",
                    fgColor="EA9999",
                ),
                "Supplier Late": PatternFill(
                    fill_type="solid",
                    fgColor="F9CB9C",
                ),
                "Waiting": PatternFill(
                    fill_type="solid",
                    fgColor="FFF2CC",
                ),
                "Received Late": PatternFill(
                    fill_type="solid",
                    fgColor="FCE5CD",
                ),
                "Received": PatternFill(
                    fill_type="solid",
                    fgColor="D9EAD3",
                ),
            }

            if status_column is not None:
                for row_number in range(
                    2,
                    worksheet.max_row + 1,
                ):
                    status_value = worksheet.cell(
                        row=row_number,
                        column=status_column,
                    ).value

                    row_fill = fills.get(status_value)

                    if row_fill is None:
                        continue

                    for column_number in range(
                        1,
                        worksheet.max_column + 1,
                    ):
                        worksheet.cell(
                            row=row_number,
                            column=column_number,
                        ).fill = row_fill

        return output_path.resolve()