from __future__ import annotations

import argparse
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.bom_engine import BomEngine
from core.bom_repository import BomRepository
from core.bom_selector import BomSelector
from core.config import OdooConfig
from core.dataset_execution_engine import DatasetExecutionEngine
from core.odoo_client import OdooClient, OdooConnectionError


QTY_TOLERANCE = 0.000001


def canon(value: Any) -> str:
    return str(value or "").strip().upper()


def qty(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


@dataclass(frozen=True)
class ComparisonKey:
    sale_line_id: int
    group_name: str
    root_sku: str
    parent_sku: str
    component_sku: str


@dataclass
class ComparisonValue:
    catalog_qty: float = 0.0
    odoo_qty: float = 0.0
    catalog_paths: set[str] | None = None
    odoo_paths: set[str] | None = None
    odoo_bom_ids: set[int] | None = None
    odoo_bom_references: set[str] | None = None

    def __post_init__(self) -> None:
        if self.catalog_paths is None:
            self.catalog_paths = set()
        if self.odoo_paths is None:
            self.odoo_paths = set()
        if self.odoo_bom_ids is None:
            self.odoo_bom_ids = set()
        if self.odoo_bom_references is None:
            self.odoo_bom_references = set()


def build_engines(
    client: OdooClient,
) -> tuple[BomEngine, DatasetExecutionEngine]:
    repository = BomRepository(client)
    selector = BomSelector(repository)
    odoo_engine = BomEngine(
        repository,
        selector,
    )

    catalog_engine = DatasetExecutionEngine(
        client,
        odoo_engine,
        environment="production",
    )

    return odoo_engine, catalog_engine


def aggregate_catalog_rows(
    rows: list[Any],
) -> dict[ComparisonKey, ComparisonValue]:
    result: dict[
        ComparisonKey,
        ComparisonValue,
    ] = {}

    for row in rows:
        if getattr(
            row,
            "source",
            "PRODUCT_CATALOG",
        ) != "PRODUCT_CATALOG":
            continue

        key = ComparisonKey(
            sale_line_id=int(
                getattr(
                    row,
                    "sale_line_id",
                    0,
                )
                or 0
            ),
            group_name=str(
                getattr(
                    row,
                    "group_name",
                    "",
                )
                or ""
            ).strip(),
            root_sku=canon(
                getattr(
                    row,
                    "root_sku",
                    "",
                )
            ),
            parent_sku=canon(
                getattr(
                    row,
                    "parent_sku",
                    "",
                )
            ),
            component_sku=canon(
                getattr(
                    row,
                    "component_sku",
                    "",
                )
            ),
        )

        value = result.setdefault(
            key,
            ComparisonValue(),
        )

        value.catalog_qty += qty(
            getattr(
                row,
                "required_qty",
                0,
            )
        )

        path = str(
            getattr(
                row,
                "path",
                "",
            )
            or ""
        ).strip()

        if path:
            value.catalog_paths.add(path)

    return result


def merge_odoo_rows(
    result: dict[ComparisonKey, ComparisonValue],
    rows: list[Any],
) -> None:
    for row in rows:
        key = ComparisonKey(
            sale_line_id=int(
                getattr(
                    row,
                    "sale_line_id",
                    0,
                )
                or 0
            ),
            group_name=str(
                getattr(
                    row,
                    "group_name",
                    "",
                )
                or ""
            ).strip(),
            root_sku=canon(
                getattr(
                    row,
                    "root_sku",
                    "",
                )
            ),
            parent_sku=canon(
                getattr(
                    row,
                    "parent_sku",
                    "",
                )
            ),
            component_sku=canon(
                getattr(
                    row,
                    "component_sku",
                    "",
                )
            ),
        )

        value = result.setdefault(
            key,
            ComparisonValue(),
        )

        value.odoo_qty += qty(
            getattr(
                row,
                "required_qty",
                0,
            )
        )

        path = str(
            getattr(
                row,
                "path",
                "",
            )
            or ""
        ).strip()

        if path:
            value.odoo_paths.add(path)

        bom_id = getattr(
            row,
            "bom_id",
            None,
        )

        if bom_id is not None:
            value.odoo_bom_ids.add(
                int(bom_id)
            )

        reference = str(
            getattr(
                row,
                "bom_reference",
                "",
            )
            or ""
        ).strip()

        if reference:
            value.odoo_bom_references.add(
                reference
            )


def status_for(
    catalog_qty: float,
    odoo_qty: float,
) -> str:
    if (
        abs(catalog_qty)
        <= QTY_TOLERANCE
        and abs(odoo_qty)
        > QTY_TOLERANCE
    ):
        return "EXTRA IN ODOO BOM"

    if (
        abs(catalog_qty)
        > QTY_TOLERANCE
        and abs(odoo_qty)
        <= QTY_TOLERANCE
    ):
        return "MISSING IN ODOO BOM"

    if abs(
        odoo_qty - catalog_qty
    ) > QTY_TOLERANCE:
        return "QTY MISMATCH"

    return "PASS"


def build_differences(
    comparison: dict[
        ComparisonKey,
        ComparisonValue,
    ]
) -> list[dict[str, Any]]:
    rows = []

    for key, value in sorted(
        comparison.items(),
        key=lambda item: (
            item[0].sale_line_id,
            item[0].group_name,
            item[0].root_sku,
            item[0].parent_sku,
            item[0].component_sku,
        ),
    ):
        status = status_for(
            value.catalog_qty,
            value.odoo_qty,
        )

        if status == "PASS":
            continue

        rows.append(
            {
                "sale_line_id": key.sale_line_id,
                "group_name": key.group_name,
                "root_sku": key.root_sku,
                "parent_sku": key.parent_sku,
                "component_sku": key.component_sku,
                "status": status,
                "catalog_qty": value.catalog_qty,
                "odoo_qty": value.odoo_qty,
                "difference": (
                    value.odoo_qty
                    - value.catalog_qty
                ),
                "odoo_bom_ids": ", ".join(
                    str(value)
                    for value in sorted(
                        value.odoo_bom_ids
                    )
                ),
                "odoo_bom_references": "; ".join(
                    sorted(
                        value.odoo_bom_references
                    )
                ),
                "catalog_paths": " | ".join(
                    sorted(
                        value.catalog_paths
                    )
                ),
                "odoo_paths": " | ".join(
                    sorted(
                        value.odoo_paths
                    )
                ),
            }
        )

    return rows


def build_parent_summary(
    differences: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    grouped: dict[
        tuple[int, str, str, str],
        list[dict[str, Any]],
    ] = defaultdict(list)

    for row in differences:
        grouped[
            (
                row["sale_line_id"],
                row["group_name"],
                row["root_sku"],
                row["parent_sku"],
            )
        ].append(
            row
        )

    summary = []

    for (
        sale_line_id,
        group_name,
        root_sku,
        parent_sku,
    ), rows in sorted(
        grouped.items()
    ):
        summary.append(
            {
                "sale_line_id": sale_line_id,
                "group_name": group_name,
                "root_sku": root_sku,
                "parent_sku": parent_sku,
                "difference_count": len(
                    rows
                ),
                "missing_count": sum(
                    row["status"]
                    == "MISSING IN ODOO BOM"
                    for row in rows
                ),
                "extra_count": sum(
                    row["status"]
                    == "EXTRA IN ODOO BOM"
                    for row in rows
                ),
                "qty_mismatch_count": sum(
                    row["status"]
                    == "QTY MISMATCH"
                    for row in rows
                ),
                "odoo_bom_ids": "; ".join(
                    sorted(
                        {
                            row[
                                "odoo_bom_ids"
                            ]
                            for row in rows
                            if row[
                                "odoo_bom_ids"
                            ]
                        }
                    )
                ),
                "odoo_bom_references": "; ".join(
                    sorted(
                        {
                            row[
                                "odoo_bom_references"
                            ]
                            for row in rows
                            if row[
                                "odoo_bom_references"
                            ]
                        }
                    )
                ),
                "status": "BOM MISMATCH",
            }
        )

    return summary


def style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    fill = PatternFill(
        "solid",
        fgColor="1F4E78",
    )

    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(
            color="FFFFFF",
            bold=True,
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for column in range(
        1,
        ws.max_column + 1,
    ):
        width = max(
            len(
                str(
                    ws.cell(
                        row=row,
                        column=column,
                    ).value
                    or ""
                )
            )
            for row in range(
                1,
                min(
                    ws.max_row,
                    500,
                ) + 1,
            )
        )

        ws.column_dimensions[
            get_column_letter(column)
        ].width = min(
            max(width + 2, 12),
            70,
        )


def write_report(
    *,
    so_number: str,
    catalog_result: Any,
    odoo_result: Any,
    differences: list[dict[str, Any]],
    parent_summary: list[dict[str, Any]],
    output_dir: Path,
) -> Path:
    output_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    timestamp = datetime.now().strftime(
        "%Y%m%d_%H%M%S"
    )

    path = (
        output_dir
        / (
            f"SO_BOM_Diagnostics_"
            f"{so_number}_{timestamp}.xlsx"
        )
    )

    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "SUMMARY"
    summary_ws.append(
        ["Metric", "Value"]
    )
    summary_ws.append(
        ["SO", so_number]
    )
    summary_ws.append(
        [
            "Dataset ID",
            getattr(
                catalog_result,
                "dataset_id",
                "",
            ),
        ]
    )
    summary_ws.append(
        [
            "Dataset Batch",
            getattr(
                catalog_result,
                "batch_reference",
                "",
            ),
        ]
    )
    summary_ws.append(
        [
            "Catalog Rows",
            len(
                getattr(
                    catalog_result,
                    "rows",
                    [],
                )
            ),
        ]
    )
    summary_ws.append(
        [
            "Odoo Rows",
            len(
                getattr(
                    odoo_result,
                    "rows",
                    [],
                )
            ),
        ]
    )
    summary_ws.append(
        [
            "BOMs with Differences",
            len(parent_summary),
        ]
    )
    summary_ws.append(
        [
            "Difference Lines",
            len(differences),
        ]
    )
    summary_ws.append(
        [
            "Catalog Fallback Count",
            len(
                getattr(
                    catalog_result,
                    "fallbacks",
                    [],
                )
                or []
            ),
        ]
    )

    parent_ws = wb.create_sheet(
        "BOM_SUMMARY"
    )
    parent_ws.append(
        [
            "Sale Line ID",
            "Group",
            "Root SKU",
            "Parent SKU",
            "Difference Count",
            "Missing",
            "Extra",
            "Qty Mismatch",
            "Odoo BOM IDs",
            "Odoo BOM References",
            "Status",
        ]
    )

    for row in parent_summary:
        parent_ws.append(
            [
                row["sale_line_id"],
                row["group_name"],
                row["root_sku"],
                row["parent_sku"],
                row["difference_count"],
                row["missing_count"],
                row["extra_count"],
                row["qty_mismatch_count"],
                row["odoo_bom_ids"],
                row[
                    "odoo_bom_references"
                ],
                row["status"],
            ]
        )

    details_ws = wb.create_sheet(
        "BOM_DIFFERENCES"
    )
    details_ws.append(
        [
            "Sale Line ID",
            "Group",
            "Root SKU",
            "Parent SKU",
            "Component SKU",
            "Status",
            "Catalog Qty",
            "Odoo Qty",
            "Odoo - Catalog",
            "Odoo BOM IDs",
            "Odoo BOM References",
            "Catalog Path",
            "Odoo Path",
        ]
    )

    for row in differences:
        details_ws.append(
            [
                row["group_name"],
                row["root_sku"],
                row["parent_sku"],
                row["component_sku"],
                row["status"],
                row["catalog_qty"],
                row["odoo_qty"],
                row["difference"],
                row["odoo_bom_ids"],
                row[
                    "odoo_bom_references"
                ],
                row["catalog_paths"],
                row["odoo_paths"],
            ]
        )

    fallback_ws = wb.create_sheet(
        "CATALOG_FALLBACKS"
    )
    fallback_ws.append(
        [
            "Sale Line ID",
            "Group",
            "Root SKU",
            "Reason",
            "Source",
        ]
    )

    for fallback in (
        getattr(
            catalog_result,
            "fallbacks",
            [],
        )
        or []
    ):
        fallback_ws.append(
            [
                getattr(
                    fallback,
                    "sale_line_id",
                    "",
                ),
                getattr(
                    fallback,
                    "group_name",
                    "",
                ),
                getattr(
                    fallback,
                    "root_sku",
                    "",
                ),
                getattr(
                    fallback,
                    "reason",
                    "",
                ),
                getattr(
                    fallback,
                    "source",
                    "",
                ),
            ]
        )

    errors_ws = wb.create_sheet(
        "ERRORS"
    )
    errors_ws.append(
        [
            "Source",
            "Sale Line ID",
            "Group",
            "Root SKU",
            "Error",
        ]
    )

    for source, result in (
        ("CATALOG", catalog_result),
        ("ODOO", odoo_result),
    ):
        for error in (
            getattr(
                result,
                "errors",
                [],
            )
            or []
        ):
            errors_ws.append(
                [
                    source,
                    getattr(
                        error,
                        "sale_line_id",
                        "",
                    ),
                    getattr(
                        error,
                        "group_name",
                        "",
                    ),
                    getattr(
                        error,
                        "root_sku",
                        "",
                    ),
                    getattr(
                        error,
                        "error",
                        "",
                    ),
                ]
            )

    for ws in (
        summary_ws,
        parent_ws,
        details_ws,
        fallback_ws,
        errors_ws,
    ):
        style_sheet(ws)

    wb.save(path)
    return path.resolve()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Palygina konkretaus SO Product Catalog "
            "BOM išskleidimą su aktyviais Odoo BOM."
        )
    )

    parser.add_argument(
        "--so-number",
        required=True,
    )

    parser.add_argument(
        "--output-dir",
        type=Path,
        default=(
            Path(__file__).resolve().parents[1]
            / "output"
        ),
    )

    return parser.parse_args()


def main() -> None:
    args = parse_args()
    so_number = str(
        args.so_number
        or ""
    ).strip()

    client = OdooClient(
        OdooConfig.from_env()
    )

    try:
        client.connect()

        (
            odoo_engine,
            catalog_engine,
        ) = build_engines(
            client
        )

        odoo_result = (
            odoo_engine.explode_sales_order(
                so_number,
                only_grouped_lines=True,
            )
        )

        catalog_result = (
            catalog_engine.explode_sales_order(
                so_number,
                only_grouped_lines=True,
            )
        )

        comparison = aggregate_catalog_rows(
            catalog_result.rows
        )

        merge_odoo_rows(
            comparison,
            odoo_result.rows,
        )

        differences = build_differences(
            comparison
        )

        parent_summary = (
            build_parent_summary(
                differences
            )
        )

        output_path = write_report(
            so_number=so_number,
            catalog_result=catalog_result,
            odoo_result=odoo_result,
            differences=differences,
            parent_summary=parent_summary,
            output_dir=args.output_dir,
        )

        print()
        print("=" * 80)
        print("SO BOM DIAGNOSTICS")
        print("=" * 80)
        print("SO:", so_number)
        print(
            "BOM su neatitikimais:",
            len(parent_summary),
        )
        print(
            "Neatitikimų eilutės:",
            len(differences),
        )
        print(
            "Catalog fallback:",
            len(
                getattr(
                    catalog_result,
                    "fallbacks",
                    [],
                )
                or []
            ),
        )
        print()
        print("Ataskaita:", output_path)

    except OdooConnectionError as exc:
        print(
            f"\nOdoo klaida: {exc}"
        )
        raise SystemExit(1) from exc
    except Exception as exc:
        print(
            f"\nNenumatyta klaida: {exc}"
        )
        raise SystemExit(1) from exc


if __name__ == "__main__":
    main()